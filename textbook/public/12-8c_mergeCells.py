#! python3
# adjustCellHeightWidth.py - Adjust cell's height and width

#--------------------------------合并单元格 merge_cells------------------------

import openpyxl
 
wb = openpyxl.Workbook()

#sheet = wb.get_active_sheet()
sheet = wb.active

sheet.merge_cells('A1:D3')
sheet['A1'] = 'Twelve cells merged together.'

sheet.merge_cells('C5:D5')
sheet['C5'] = 'Two merged cells.'

wb.save('merged.xlsx')


""" 解除merge
sheet.unmerge_cells('A1:D3')
sheet.unmerge_cells('C5:D5')
wb.save('merged.xlsx')
"""