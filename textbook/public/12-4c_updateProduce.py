#! python3
# updateProduce.py - Corrects costs in produce sales spreadsheet.
#——----------------------遍历单元格更改excel数据-----------------------
import openpyxl

wb = openpyxl.load_workbook('produceSales.xlsx')
print(wb.sheetnames)
sheet = wb['Sheet']

# 需要更改的数据存储在字典中
PRICE_UPDATES = {'Garlic': 3.07,
                 'Celery': 1.19,
                 'Lemon': 1.27,
                 'Red onion': 0.77
}

# 遍历Excel单元格
for rowNum in range(2, sheet.max_row+1): # 第一行是index 所以从第二行开始
    produceName = sheet.cell(row=rowNum, column=1).value #取出excel栏位名
    if produceName in PRICE_UPDATES:   #对比key值和excel栏位
        # Assign the updated price of produceName to cell(row=rowNum, column=2)
        sheet.cell(row=rowNum, column=2).value = PRICE_UPDATES[produceName]

wb.save('updatedProduceSales.xlsx')
