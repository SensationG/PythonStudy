import openpyxl
from openpyxl.utils.cell import get_column_letter, column_index_from_string
#------------------------字母《--》数字 转换----------------
print(get_column_letter(1))
#'A'

print(get_column_letter(2))
#'B'

print(get_column_letter(27))
#'AA'

print(get_column_letter(900))
#'AHP'

print(column_index_from_string('A'))
#1

print(column_index_from_string('AA'))
#27