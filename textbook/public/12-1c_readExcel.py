import openpyxl
#---------------读取/遍历/属性/excel------------------------

# Load in the workbook 加载Excel表格

wb = openpyxl.load_workbook('example.xlsx')

# 1 Get sheet names
a=wb.get_sheet_names()  #获取所有工作表名称
print(a)
print(wb.sheetnames)  #打印excel中的所有工作表名称

# 2 Get a sheet by name  #通过工作表名字获取工作表
# wb.get_sheet_by_name('Sheet3') 
sheet = wb['Sheet3'] #赋值后可对工作表进行操作
print(sheet.title) 

# 3 Get active sheet  获取打开时候显示的表格
# wb.get_active_sheet()
sheet = wb.active
print("active sheet: " + sheet.title)

# 4 Getting Cells from the Sheets 查看参数的属性和方法列表
#print(dir(sheet))

# 5 Retrieve the maximum amount of rows 获取表格的最大行数和列数
# old: sheet.get_highest_row() 
# sheet.max_row

# Retrieve the maximum amount of columns
# old: sheet.get_highet_column() 
# sheet.max_row/sheet.max_column 

print("This sheet has " + str(sheet.max_row) + ' rows ' + str(sheet.max_column) + ' columns.')

# 6 Retrieve the value of a certain cell 获取某个单元格的value,所在行列
print(sheet['A1'].value)  #赋值时直接赋值即可不需要value

# Select element 'B1' of your sheet c = sheet["%s%d" % ("B", 1)]
c = sheet['B1']
print(c.value)  #'Apples'
# 打印所在行列
print('Row ' + str(c.row) + ', Column ' + str(c.column) + ' is ' + c.value)
# c.coordinate 打印c的坐标
print('Cell ' + c.coordinate + ' is ' + c.value)
# 'Cell B1 is Apples'

# 7 遍历 按列 
for i in range(1, 8, 2):
  print(i, sheet.cell(row=i, column=2).value)

# 8 获取区域数据 sheet['A1':'C3'] 实际上也是按行读取

for rows in sheet['A1':'C3']:
  for cellObj in rows:
    print(cellObj.coordinate, cellObj.value)
  print('---- END OF ROW ----')
