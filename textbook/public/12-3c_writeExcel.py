import openpyxl
#--------------------------创建excel/新建分页------------
# 1 Creating a new workbook 创建一个excel文件
wb = openpyxl.Workbook()
print(wb.sheetnames)

# 2 Create one sheet 创建一个新的分页 create_sheet(title)
wb.create_sheet()
wb.create_sheet('a')

print("After creating one Sheet" )
print(wb.sheetnames)

# 3 创建分页并选择分页创建的位置index/title
print("After creating 'First Sheet' and 'Middle Sheet'" )#创建分页并选择分页位置
wb.create_sheet(index=0, title='First Sheet')
wb.create_sheet(index=2, title='Middle Sheet')

print(wb.sheetnames)

# 4 remove_sheet() takes a Worksheet object, not a string of the sheet name, as its argument.
wb.remove(wb['Middle Sheet']) #删除分页
wb.remove(wb['Sheet1'])
print("After removing 'Middle Sheet' and 'Sheet1'")
print(wb.sheetnames)

# 5 取打开excel时显示的那页并赋值
sheet = wb.active
print("active sheet:" + sheet.title)
sheet['A1'] = 'Hello'
print(sheet['A1'].value)