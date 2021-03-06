from openpyxl import Workbook
from openpyxl.utils.cell import get_column_letter, column_index_from_string
#------------------------将二维list数据写入excel 并带有数据标题栏----------------------------------

# initialize data
grades_dist = [['30-39', 2], ['40-49', 8], ['50-59', 10], ['60-69', 7]]
labels = ['grade range', 'count']
start_row = 2 #写入的起始位置
start_col = 13
col_letter = get_column_letter(start_col)

# Create new workbook
book = Workbook()
sheet = book.active

# Set the width of the column start_col 
sheet.column_dimensions[col_letter].width = len(labels[0])+2

# Write column labels 写入数据标题
for i, label in enumerate(labels):
    sheet.cell(row=1, column=start_col+i, value=label)
    
# Write data to row, i: row index, j: column index 写入数据
for i, item in enumerate(grades_dist):
    for j, value in enumerate(item):
        sheet.cell(row=2+i, column=start_col+j, value=value)

print("done")
book.save("range_written.xlsx")

