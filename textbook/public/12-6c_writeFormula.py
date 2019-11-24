#! python3
# updateProduce.py - Corrects costs in produce sales spreadsheet.

#-------------------------通过python使用excel内置公式--------------
import openpyxl

wb = openpyxl.Workbook()

sh = wb.active
sd = wb.create_sheet('a') #新建一个分页

sh['A1'] = 200
sh['A2'] = 300
sh['A3'] = '=SUM(A1:A2)'#可以直接用excel的公式（载入excel时设置：data_only=False)

sd['B1'] = 399
sd['B2'] = 300
sd['B3'] = '=SUM(B1:B2)'

wb.save('writeFormula.xlsx')

print(sh['A3'].value)

wbDataOnly = openpyxl.load_workbook('writeFormula.xlsx', data_only=Ture)#想要执行excel公式要设为True

sh1 = wbDataOnly.active

#openpyxl doesn't evaluate formula, it returns None.
print(sh1['A3'].value)

