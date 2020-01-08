# For each single excel file, create one CSV file per sheet. 
# The filenames of the CSV files should be <excel filename>_<sheet title>.csv,
#-------------------------读取excel写入到csv---每各工作簿单独生成一个csv文件----------------------------
import os
import csv
import openpyxl

for filename in os.listdir('.'):
    
    #查找是xlsx的excel文件
    if not filename.endswith('.xlsx'):
        continue # skip non-csv files
    
    fname, ext = os.path.splitext(filename) #分离文件名/文件类型
    
    # open file using openpyxl
    wb = openpyxl.load_workbook(filename) #加载该excel文件

    for sh in wb.sheetnames: #循环所有工作簿
    # Loop through every sheet in the workbook.
        sheet = wb[sh]
        #每个工作簿单独生成一个csv文件
        # Create the CSV filename from the Excel filename and sheet title.
        # Create the csv.writer object for this CSV file.
        csv_fname = fname + '_'  + sh + '.csv' #设定转换后的档名
        print(csv_fname)
        
        cf = open(csv_fname, 'w', newline='') #新建csv档
        csv_writer = csv.writer(cf) 
        
        # iter_rows按行遍历excle 写入csv
        for rowdata in sheet.iter_rows(min_row=1, values_only=True):
            csv_writer.writerow(rowdata) #写入
        
        cf.close()