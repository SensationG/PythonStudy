# -*- coding: utf-8 -*-
"""
Created on Sun Nov 24 20:27:42 2019

@author: hhw
"""

import openpyxl

'''
写入
'''
# 1 新建Excel表格
wb = openpyxl.Workbook()
# 2 新建分页
sheet = wb.active

# 3 使用二维list/配合append写入数据 
rows = [
    [88, 46, '中文'],
    (89, 38, 12),
    (23, 59, 78),
    (56, 21, 98),
    (24, 18, 43),
    (34, 15, 67)
]
for row in rows:
    sheet.append(row) #每次写入一行
    
# 4 使用value 遍历赋值写入 **list 中的数据（添加数据标题为例）
start_column=1
labels = ['grade range','count']   #写入标题（标准写法）
for i, label in enumerate(labels):
     sheet.cell(row=1, column=start_column+i, value=label)    
     
# 5 使用value 遍历写入 **字典 中的数据(key+value)
g_dict={'t1':21,'t2':78,'t3':311,'t4':213,'t5':213}
for i,v in enumerate(g_dict.items()):
     for j,kv in enumerate(v):
         sheet.cell(row=2+i,column=start_column+j,value=kv)
         #sheet.cell(row=2+i,column=start_column+j).value=kv 第二种写法
# 6 保存
wb.save('example.xlsx')

'''
读取
'''
# 1 读取Excel表格
wb = openpyxl.load_workbook('example.xlsx')
# 2 新建Excel表格 (注意Workbook首字母要大写)
#wb = openpyxl.Workbook()
# 3 查看Excel所有分页名
print(wb.sheetnames)
# 4 新建Excel分页
wb.create_sheet('sheet1') #wb.create_sheet(index=2, title='Middle Sheet') 可选择位置
# 5 删除分页
#wb.remove(wb['Sheet1'])
# 6 给分页赋值以方便操作
sheet = wb.active #直接赋值excel启动后首先显示的分页
sheet1 = wb['sheet1'] #赋值指定名称分页
# 7 获取表格的最大行/列数 row:行  column:列
print('-7')
print(sheet.max_row,sheet.max_column)
# 8 获取单个单元格的信息（值，所在行，列）
print('-8')
c = sheet['B1']
print(c.value,c.row,c.column)  
# 9 遍历所有单元格 (按区域遍历)
print('-9')
for rows in sheet['A1':'C3']: #每次选取一行
  for cellObj in rows: #每行种的每一个数据遍历
    print(cellObj.coordinate, cellObj.value) #coordinate为列名
  
# 10 遍历所有单元格 (按表格长度/宽度) 可配合max_row/max_column
print('-10')
for i in range(1,sheet.max_row):
    print(i)
    for j in range(1,sheet.max_column):
        print(sheet.cell(row=i, column=j).value) #取出指定单元格数据
        
# 11 快速读取某行的信息 (也可以扩展成按行读取信息)
print('-11')
first_row = list(sheet.rows)[0]
for i in first_row:
    print(i.value)
    
# 12 快速按行读取信息并进行数据操作
print('-12')
for row in sheet.iter_rows(min_row=1,min_col=1,max_row=sheet.max_row,max_col=sheet.max_column):
    data=[]
    #data=[cell.value for cell in row] #快速添加数据
    for cell in row: #每行的每个数据
        print(cell.value, end=" ")
        data.append(cell.value) #添加每行的每个数据
    print()
    #print(data)
   
'''
绘图
'''   
from openpyxl.chart import (
    Reference,
    Series,
    BarChart
)
wb = openpyxl.load_workbook('example.xlsx')
sheet = wb.active
# 1 绘制直方图 x轴数据+y轴数据
data = Reference(sheet, min_col=2, min_row=1, max_col=2, max_row=5) #y轴数据
categs = Reference(sheet, min_col=1, min_row=2, max_col=1,max_row=5) #x轴数据

chart = BarChart() #直方图类型
chart.add_data(data, titles_from_data=True) #添加y轴数据(数据源包含标题 所以设true)
chart.set_categories(categs) #添加x轴数据

chart.varyColors = True #彩色
chart.title = "Olympic Gold medals in London"
chart.x_axis.title = 'Rank'    #x轴名称
chart.y_axis.title = 'Value'   #y轴名称
sheet.add_chart(chart, "E2")  #存放位置
chart.legend = None #不需要图例

wb.save("example.xlsx")














