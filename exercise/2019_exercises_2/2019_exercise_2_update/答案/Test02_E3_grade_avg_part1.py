# -*- Coding: utf-8 -*-

import os
from openpyxl import load_workbook, Workbook
from openpyxl.utils.cell import get_column_letter, column_index_from_string
from openpyxl.chart import Reference, BarChart
from datetime import datetime

'''Predict a file's encoding using chardet'''
#---------------------检测txt文件编码格式并读取------------------------
def predict_encoding(file_path):
    import chardet
    #import os
    encoding = ''
    bytes = min(1024, os.path.getsize(file_path))

    # Open the file as binary data
    raw = open(file_path, 'rb').read(bytes)

    result = chardet.detect(raw)
    encoding = result['encoding']
    #print("original encoding:", encoding)
    if not encoding or encoding in ['Big5', 'cp950','ISO-8859-1']:
        encoding = None
    
    return encoding

# Read credits.txt and create credits dic
folder = "data"
filename = 'credits.txt'
filepath = os.path.join(folder, filename)

credits_dict = {} # {course: credits}

with open(filepath, encoding=predict_encoding(filepath)) as f: 
    for line in f:
        words = line.strip().split(":")
        k = words[0].strip()
        v = words[1].strip()
        credits_dict[k] = int(v)

print(credits_dict)

#-----------------------------打开excel档案--------------------------------------
# Read grades file
xlsxfile = "grades.xlsx"
xlsxpath = os.path.join(folder, xlsxfile)

# Read workbook
wb = load_workbook(xlsxpath)

# Get grade sheet
sheet = wb['工作表1']

# Get first row values only
first_row = list(sheet.values)[0]
credits = []
# credits = [credits_dict[c] for c in first_row[1:]]

# Get courses in first row
print("First row:")
for c in first_row[1:]:
    print(c, credits_dict[c], end=', ')
    credits.append(credits_dict[c])
print("\n")

# Read data row by row
data_rows  = [] # [[sid1, avg1], [sid2, avg2],...]
avgs = []       # [avg1, avg2, ...]
print("Data row by row:")
##----------------------累计分数-----------------------------------------------
#for row in sheet.iter_rows(min_row=2, min_col=1, max_row=sheet.max_row, max_col=sheet.max_column):
for row in sheet.iter_rows(min_row=2, values_only=True):
    #print(row)
    sid = row[0]
    total_scores = 0
    total_credits = 0 
    for i, value in enumerate(row[1:]):
        print('{0:.2f}: {1}'.format(value, credits[i]))
        total_scores +=  int(value) * credits[i]
        total_credits += credits[i]
    print(total_credits)
    
    avg = total_scores/total_credits
    
    #print('{0}: {1:.2f}'.format(sid, avg))
    # Save [sid, avg] in data_rows
    data_rows.append([sid, round(avg, 2)])
    avgs.append(avg)
    print()

print(data_rows)

# Write to a new file


# Grouping avgs


# Write group data to sheet    


# Create a bar chart using group data
