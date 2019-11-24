
# coding='utf-8'

from openpyxl import Workbook

book = Workbook()
sheet = book.active

rows = [
    [88, 46, '中文'],
    (89, 38, 12),
    (23, 59, 78),
    (56, 21, 98),
    (24, 18, 43),
    (34, 15, 67)
]

# Write data to sheet row by row
for row in rows:
    sheet.append(row)

# 讀取已寫入的 data

# 方法 1: Get rows
# Get first row
first_row = list(sheet.rows)[0]

# Get its values
print("First row:")
for c in first_row:
    print(c.value, end= ' ')
print("\n")
    
# Read data row by row
print("Data row by row:")
for row in sheet.iter_rows(min_row=1, min_col=1, max_row=6, max_col=3):
    #print(row)
    data = []
    #data = [cell.value for cell in row]
    for cell in row:
        print(cell.value, end=" ")
        data.append(cell.value)
    print(" --> data in list:", data)
    print()


# 方法 2: Get values only
# Get first row's value
first_row = list(sheet.values)[0]

# Get its values
print("First row:")
for value in first_row:
    print(value, end= ' ')
print("\n")
    
# Read data row by row
print("Data row by row:")
for row in sheet.iter_rows(min_row=1, min_col=1, values_only=True):
    #print(row)
    data = []
    #data = [value for value in row]
    for value in row:
        print(value, end=" ")
        data.append(value)
    print(" --> data in list:", data)
    print()
    
book.save('appending.xlsx')