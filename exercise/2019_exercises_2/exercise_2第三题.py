# -*- coding: utf-8 -*-
"""
Created on Thu Nov 21 09:18:34 2019

@author: hhw
"""

#exercise2-3
import os
import pandas as pd
from datetime import datetime, timedelta
import openpyxl
from openpyxl.chart import BarChart, Reference

#读取修课成绩
df_g = pd.read_excel("data/grades.xlsx")
print(df_g.head(3))
print(df_g.shape)

#读取开课学分数
courages={} 
with open('data/credits.txt', 'r', encoding='utf=8') as r:
    for line in r:
        cour=line.split(':')[0].strip()
        cred=line.split(':')[1].strip()
        courages[cour]=cred        
#print(courages)
 
# 1 计算修课平均成绩
#print(df_g.iloc[0][2])#取dataframe的某行某列
s_avg={}
for i in range(len(df_g)):  #遍历df行列
    sum=0
    for j in range(1,df_g.shape[1]):
        sum+=df_g.iloc[i][j]
    #print(sum)
    avg=sum/5
    s_avg[df_g.iloc[i][0]]=avg
print(s_avg)

#创建excel/日期
dt = datetime.now()
date1=dt.date() #日期

wb = openpyxl.Workbook() #创建工作表
wb.create_sheet('平均成绩') #创建分页
print(wb.sheetnames)
c = wb['平均成绩']

#每個人的平均分数据写入表格/遍历s_avg={}/統計平均分數段人數
i=1
grade_count={} #統計平均分數段人數
grade_count.setdefault('50-59',0)
grade_count.setdefault('60-69',0)
grade_count.setdefault('70-79',0)
grade_count.setdefault('80-89',0)
grade_count.setdefault('90-99',0)
for k,v in s_avg.items():
    j=1
    #print(k,v)
    if v>=50 and v<60:
        grade_count['50-59']+=1
    elif v>=60 and v<70:
        grade_count['60-69']+=1
    elif v>=70 and v<80:
        grade_count['70-79']+=1
    elif v>=80 and v<90:
        grade_count['80-89']+=1
    elif v>=90 and v<100:
        grade_count['90-99']+=1
    c.cell(row=i, column=j).value = k
    c.cell(row=i, column=j+1).value = v
    i+=1
print(grade_count)

#寫入平均分數段到表格
start_column=12
labels = ['grade range','count']   #写入标题（标准写法）
for i, label in enumerate(labels):
    c.cell(row=1, column=start_column+i, value=label)
#c.cell(row=1, column=12).value = 'grade range'
#c.cell(row=1, column=13).value = 'count'
    
 #从字典写入数据(标准写法)
for i,v in enumerate(grade_count.items()):
    #print(i,v)
    #print(type(v)) 数据类型 tuple
    for j,kv in enumerate(v): #j是key value是值
        #print(kv) #包含标题+value的值
        c.cell(row=i+2, column=start_column+j).value = kv #因为i从0开始 而数据从第二行开始写 所以i+2
 
    
#畫圖
# Data series is a reference to cell ranges 设定要绘图数据
data = Reference(c, min_col=13, min_row=2, max_col=13, max_row=6)
labels = Reference(c, min_col=12, min_row=2, max_col=12, max_row=6)
# Create a barChart 创建图表
chart = BarChart()

# Add the data series to the chart
chart.add_data(data)  #设定y轴
chart.set_categories(labels) #设定x轴

# set the chart's title and legend #设定画图参数
chart.title = "學生平均成績"
chart.x_axis.title = 'Grade Range'
chart.y_axis.title = 'Count'
chart.legend = None
chart.varyColors = True

# anchor the top-lfet corner of chart at E2 图放的位置
c.add_chart(chart, "D2")    
    
#保存表格
wb.save('data/grades_'+str(date1)+'.xlsx') #存储在指定位置

    
